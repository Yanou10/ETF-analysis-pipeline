"""
suivi_etf.py — Pipeline local de suivi hebdomadaire d'un panier d'ETF (Phase 1).

Étapes orchestrées par main() :
  1. Récupération de l'historique de fin de journée de chaque ETF (Yahoo Finance).
  2. Calcul d'indicateurs simples selon des règles fixes (variation sur
     ~5 séances, moyenne mobile, alerte « sous la moyenne mobile »).
  3. Écriture d'un instantané dans un fichier Excel (une feuille datée par run).
  4. Génération d'un résumé FACTUEL et neutre via Claude (Amazon Bedrock,
     API Converse).

AVERTISSEMENT : ce projet est une aide à la décision et un exercice technique.
Les indicateurs calculés sont des alertes mécaniques, PAS des ordres ni des
conseils d'achat ou de vente.
"""

import os
from datetime import datetime, timedelta

import pandas as pd
import requests

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
# Liste des symboles ETF au format Yahoo Finance (ex. « SPY »). Surchargeable via
# la variable d'environnement SYMBOLES_ETF (valeurs séparées par des virgules),
# ce qui permet de partager la même configuration entre le script local et la
# fonction Lambda.
SYMBOLES_ETF = [
    symbole.strip()
    for symbole in os.environ.get("SYMBOLES_ETF", "SPY,QQQ,VTI").split(",")
    if symbole.strip()
]

# Fenêtre de la moyenne mobile, exprimée en nombre de séances.
FENETRE_MM = 20

# Nombre de séances utilisées pour calculer la variation récente (« ~5 séances »).
FENETRE_VARIATION = 5

# Profondeur d'historique demandée à l'API, en jours calendaires. 200 jours
# couvrent largement une moyenne mobile sur 20 séances (week-ends/fériés inclus).
JOURS_HISTORIQUE = 200

# Région AWS. En environnement Lambda, AWS_REGION est défini automatiquement ;
# en local, boto3 utilise la région configurée via « aws configure » si la
# variable n'est pas présente.
REGION_AWS = os.environ.get("AWS_REGION") or "us-east-1"

# Identifiant du modèle Bedrock (Claude Haiku 4.5).
# REMARQUE : l'identifiant exact dépend de la région et nécessite le plus souvent
# un préfixe de profil d'inférence (ex. « us. » en us-east-1, « eu. » en Europe).
# À VÉRIFIER dans la console Amazon Bedrock de votre région avant utilisation.
MODELE_BEDROCK = os.environ.get(
    "MODELE_BEDROCK", "us.anthropic.claude-haiku-4-5-20251001-v1:0"
)

# Chemin du fichier Excel de suivi généré en local.
CHEMIN_EXCEL = os.environ.get("CHEMIN_EXCEL", "suivi-etf.xlsx")


# ---------------------------------------------------------------------------
# Récupération des données
# ---------------------------------------------------------------------------
def fetch_history(symbole):
    """Récupère l'historique de clôture d'un ETF via l'API chart de Yahoo Finance.

    L'endpoint public de Yahoo Finance ne nécessite aucune clé API. Un en-tête
    « User-Agent » de navigateur est envoyé afin de réduire le risque de blocage.

    Args:
        symbole (str): symbole de l'ETF au format Yahoo Finance (ex. « SPY »).

    Returns:
        pandas.DataFrame: indexé par date (croissante), avec au minimum une
        colonne numérique « close ».

    Raises:
        RuntimeError: si Yahoo Finance ne renvoie aucune donnée exploitable.
        requests.HTTPError: si l'API renvoie un statut d'erreur.
    """
    fin = int(datetime.now().timestamp())
    debut = int((datetime.now() - timedelta(days=JOURS_HISTORIQUE)).timestamp())
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbole}"
    parametres = {
        "period1": debut,
        "period2": fin,
        "interval": "1d",   # données quotidiennes
    }
    # Yahoo bloque les requêtes sans User-Agent : on imite un navigateur.
    entetes = {"User-Agent": "Mozilla/5.0"}

    reponse = requests.get(url, params=parametres, headers=entetes, timeout=30)
    reponse.raise_for_status()
    donnees = reponse.json()

    # Structure attendue : chart.result[0].{timestamp, indicators.quote[0].close}.
    resultat = (donnees.get("chart") or {}).get("result")
    if not resultat:
        raise RuntimeError(f"Aucune donnée renvoyée par Yahoo Finance pour {symbole}.")

    bloc = resultat[0]
    horodatages = bloc.get("timestamp")
    quote = (bloc.get("indicators", {}).get("quote") or [{}])[0]
    cloture = quote.get("close")
    if not horodatages or not cloture:
        raise RuntimeError(f"Données de clôture indisponibles pour {symbole}.")

    # Construction du DataFrame : date en index, colonne « close » numérique.
    df = pd.DataFrame({
        "date": pd.to_datetime(horodatages, unit="s"),
        "close": cloture,
    })
    df = df.set_index("date").sort_index()
    df["close"] = pd.to_numeric(df["close"], errors="coerce")
    df = df.dropna(subset=["close"])

    if df.empty:
        raise RuntimeError(f"Aucun cours de clôture exploitable pour {symbole}.")
    return df


# ---------------------------------------------------------------------------
# Calcul des indicateurs
# ---------------------------------------------------------------------------
def compute_indicators(symbole, df):
    """Calcule les indicateurs de suivi pour un ETF.

    Indicateurs (règles fixes, ce sont des ALERTES, pas des ordres) :
      - variation en pourcentage sur ~5 séances (FENETRE_VARIATION) ;
      - moyenne mobile sur FENETRE_MM séances ;
      - booléen « sous la moyenne mobile » : dernier prix < dernière moyenne.

    Args:
        symbole (str): symbole de l'ETF.
        df (pandas.DataFrame): historique indexé par date croissante, contenant
            une colonne numérique « close ».

    Returns:
        dict: instantané des indicateurs pour ce symbole.
    """
    df = df.sort_index()
    cours = df["close"]

    # Moyenne mobile (min_periods=1 pour rester défini même sur un court historique).
    moyenne_mobile = cours.rolling(window=FENETRE_MM, min_periods=1).mean()

    dernier_prix = float(cours.iloc[-1])
    derniere_mm = float(moyenne_mobile.iloc[-1])

    # Variation sur ~5 séances : on compare le dernier cours à celui d'il y a
    # FENETRE_VARIATION séances. None si l'historique est trop court.
    if len(cours) > FENETRE_VARIATION:
        prix_reference = float(cours.iloc[-(FENETRE_VARIATION + 1)])
        variation_pct = (dernier_prix / prix_reference - 1.0) * 100.0
    else:
        variation_pct = None

    # Alerte « sous la moyenne mobile ».
    sous_moyenne_mobile = dernier_prix < derniere_mm

    # Date de la dernière observation, au format texte pour le rapport/Excel.
    date_observation = pd.Timestamp(df.index[-1]).strftime("%Y-%m-%d")

    return {
        "symbole": symbole,
        "date": date_observation,
        "dernier_prix": dernier_prix,
        "moyenne_mobile": derniere_mm,
        "variation_5j_pct": variation_pct,
        "sous_moyenne_mobile": sous_moyenne_mobile,
    }


# ---------------------------------------------------------------------------
# Écriture Excel
# ---------------------------------------------------------------------------
def write_excel(snapshot, chemin):
    """Écrit l'instantané dans un fichier Excel : une feuille datée par exécution.

    Si le fichier existe déjà, une nouvelle feuille datée y est ajoutée afin de
    conserver l'historique des exécutions ; sinon le fichier est créé.

    Args:
        snapshot (list[dict]): liste d'indicateurs renvoyés par compute_indicators.
        chemin (str): chemin du fichier Excel à créer ou compléter.
    """
    import openpyxl  # import local : openpyxl n'est nécessaire que pour l'Excel

    if os.path.exists(chemin):
        classeur = openpyxl.load_workbook(chemin)
    else:
        classeur = openpyxl.Workbook()
        # On retire la feuille vide créée par défaut à la construction.
        classeur.remove(classeur.active)

    # Nom de feuille unique basé sur l'horodatage (caractères valides pour Excel).
    nom_base = datetime.now().strftime("%Y-%m-%d_%H%M")
    nom_feuille = nom_base
    suffixe = 1
    while nom_feuille in classeur.sheetnames:
        nom_feuille = f"{nom_base}_{suffixe}"
        suffixe += 1

    feuille = classeur.create_sheet(title=nom_feuille)

    # Ligne d'en-têtes.
    entetes = [
        "Symbole",
        "Date",
        "Dernier prix",
        f"Moyenne mobile ({FENETRE_MM} séances)",
        f"Variation {FENETRE_VARIATION} séances (%)",
        "Sous la moyenne mobile",
    ]
    feuille.append(entetes)

    # Lignes de données.
    for entree in snapshot:
        variation = entree["variation_5j_pct"]
        feuille.append(
            [
                entree["symbole"],
                entree["date"],
                round(entree["dernier_prix"], 4),
                round(entree["moyenne_mobile"], 4),
                round(variation, 2) if variation is not None else "N/A",
                "Oui" if entree["sous_moyenne_mobile"] else "Non",
            ]
        )

    classeur.save(chemin)


# ---------------------------------------------------------------------------
# Résumé en langage naturel via Claude (Amazon Bedrock, API Converse)
# ---------------------------------------------------------------------------
def summarize_with_claude(snapshot):
    """Génère un résumé factuel et neutre des mouvements via Claude (Bedrock).

    Le system prompt impose un résumé strictement descriptif : aucune
    recommandation, aucun conseil d'achat/vente, aucune prévision.

    Args:
        snapshot (list[dict]): indicateurs renvoyés par compute_indicators.

    Returns:
        str: texte du résumé produit par le modèle.
    """
    import boto3  # import local : boto3 n'est requis que pour l'appel Bedrock

    client = boto3.client("bedrock-runtime", region_name=REGION_AWS)

    # Mise en forme des données factuelles transmises au modèle.
    lignes = []
    for entree in snapshot:
        variation = entree["variation_5j_pct"]
        variation_txt = f"{variation:+.2f} %" if variation is not None else "indisponible"
        position_mm = (
            "sous la moyenne mobile"
            if entree["sous_moyenne_mobile"]
            else "au-dessus de la moyenne mobile"
        )
        lignes.append(
            f"- {entree['symbole']} (au {entree['date']}) : "
            f"dernier prix {entree['dernier_prix']:.2f}, "
            f"moyenne mobile {entree['moyenne_mobile']:.2f}, "
            f"variation sur {FENETRE_VARIATION} séances {variation_txt}, "
            f"{position_mm}."
        )
    donnees_factuelles = "\n".join(lignes)

    # System prompt : cadre strictement descriptif et neutre.
    invite_systeme = [
        {
            "text": (
                "Tu es un assistant qui résume des données de marché de manière "
                "strictement factuelle et neutre, en français. Décris uniquement "
                "les chiffres fournis (prix, moyennes mobiles, variations, position "
                "par rapport à la moyenne mobile). N'émets AUCUNE recommandation, "
                "AUCUN conseil d'achat ou de vente, AUCUNE prévision et AUCUN "
                "jugement de valeur. Reste descriptif et concis."
            )
        }
    ]

    messages = [
        {
            "role": "user",
            "content": [
                {
                    "text": (
                        "Voici les indicateurs de la semaine pour un panier d'ETF. "
                        "Rédige un résumé factuel des mouvements observés :\n\n"
                        f"{donnees_factuelles}"
                    )
                }
            ],
        }
    ]

    # Appel de l'API Converse de Bedrock.
    reponse = client.converse(
        modelId=MODELE_BEDROCK,
        messages=messages,
        system=invite_systeme,
        inferenceConfig={"maxTokens": 600, "temperature": 0.2},
    )

    return reponse["output"]["message"]["content"][0]["text"]


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------
def main():
    """Orchestre la récupération, le calcul, l'écriture Excel et le résumé."""
    print(f"Suivi de {len(SYMBOLES_ETF)} ETF : {', '.join(SYMBOLES_ETF)}")

    # 1 & 2. Récupération de l'historique et calcul des indicateurs.
    snapshot = []
    for symbole in SYMBOLES_ETF:
        print(f"  → Récupération et calcul pour {symbole}...")
        df = fetch_history(symbole)
        snapshot.append(compute_indicators(symbole, df))

    # 3. Écriture du fichier Excel de suivi.
    write_excel(snapshot, CHEMIN_EXCEL)
    print(f"Instantané écrit dans « {CHEMIN_EXCEL} ».")

    # 4. Résumé en langage naturel via Claude.
    try:
        resume = summarize_with_claude(snapshot)
        print("\n===== Résumé de la semaine =====\n")
        print(resume)
    except Exception as erreur:  # pragma: no cover
        # On n'interrompt pas l'exécution si Bedrock est inaccessible : le
        # fichier Excel a déjà été écrit. On informe simplement l'utilisateur.
        print(
            "\n[Avertissement] Le résumé via Bedrock n'a pas pu être généré "
            f"(vérifiez l'accès au modèle et la configuration AWS) : {erreur}"
        )


